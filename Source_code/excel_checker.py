import openpyxl

def analyze_excel(file_path, level):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb.active
    except Exception as e:
        return f"❌ Error opening file: {e}"

    report = [f"📊 Checking Excel file for {level} scenario..."]
    checks = []

    def col_has_formulas(col_letter, keyword):
        """Check if at least N rows in a column contain the expected formula."""
        formulas = [
            str(ws[f"{col_letter}{row}"].value).upper()
            for row in range(2, ws.max_row + 1)
            if ws[f"{col_letter}{row}"].value
        ]
        return all(keyword in f for f in formulas if f)

    if level == "Beginner":
        # SUM column (E)
        has_sum = col_has_formulas("E", "SUM(")
        report.append("✔️ SUM formulas found in column E" if has_sum else "❌ SUM formulas missing/incomplete")
        checks.append(has_sum)

        # AVERAGE column (F)
        has_avg = col_has_formulas("F", "AVERAGE(")
        report.append("✔️ AVERAGE formulas found in column F" if has_avg else "❌ AVERAGE formulas missing/incomplete")
        checks.append(has_avg)

        # Conditional formatting applied
        has_cf = len(ws.conditional_formatting) > 0
        report.append("✔️ Conditional formatting detected" if has_cf else "❌ Conditional formatting missing")
        checks.append(has_cf)

    elif level == "Intermediate":
        # Total Sales column (E) → must be Price*Quantity
        formulas = [
            str(ws[f"E{row}"].value).upper()
            for row in range(2, ws.max_row + 1)
            if ws[f"E{row}"].value
        ]
        has_formula = all("*" in f for f in formulas if f)
        report.append("✔️ Total Sales formulas applied consistently" if has_formula else "❌ Total Sales formulas incorrect/missing")
        checks.append(has_formula)

        # Lookup (Manager column F)
        lookup_formulas = [
            str(ws[f"F{row}"].value).upper()
            for row in range(2, ws.max_row + 1)
            if ws[f"F{row}"].value
        ]
        lookup_found = all(("VLOOKUP(" in f or "XLOOKUP(" in f) for f in lookup_formulas if f)
        report.append("✔️ Lookup formulas found in Manager column" if lookup_found else "❌ Lookup missing/incomplete")
        checks.append(lookup_found)

        # Data validation
        has_dv = any(ws.data_validations.dataValidation)
        report.append("✔️ Data validation dropdown exists" if has_dv else "❌ No dropdowns found")
        checks.append(has_dv)

        # Pivot table sheet
        pivot_detected = any("pivot" in s.lower() for s in wb.sheetnames)
        report.append("✔️ Pivot sheet found" if pivot_detected else "⚠️ No pivot sheet detected")
        checks.append(pivot_detected)

    elif level == "Advanced":
        # IF formulas in Result column (C)
        if_formulas = [
            str(ws[f"C{row}"].value).upper()
            for row in range(2, ws.max_row + 1)
            if ws[f"C{row}"].value
        ]
        if_formula = all("IF(" in f for f in if_formulas if f)
        report.append("✔️ IF formulas correct in Result column" if if_formula else "❌ IF formula missing/incomplete")
        checks.append(if_formula)

        # Pivot table sheet
        pivot_detected = any("pivot" in s.lower() for s in wb.sheetnames)
        report.append("✔️ Pivot sheet found" if pivot_detected else "⚠️ No pivot sheet detected")
        checks.append(pivot_detected)

        # Chart detection
        has_chart = hasattr(ws, "_charts") and bool(ws._charts)
        report.append("✔️ Chart present" if has_chart else "❌ Chart missing")
        checks.append(has_chart)

        # Macro detection
        has_macro = getattr(wb, "vba_archive", None) is not None
        report.append("✔️ Macro detected" if has_macro else "❌ Macro missing")
        checks.append(has_macro)

    else:
        return "❌ Invalid proficiency level."

    # === Final score ===
    total = len(checks)
    correct = sum(checks)
    score_percent = round((correct / total) * 100, 1) if total > 0 else 0
    result = "✅ PASS" if score_percent >= 70 else "❌ FAIL"

    report.append(f"\n📊 Final Score: {correct}/{total} ({score_percent}%) → {result}")
    return "\n".join(report)
